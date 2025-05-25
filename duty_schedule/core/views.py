from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth import authenticate, login
from django.utils.timezone import now
from datetime import datetime
from .forms import LoginForm, RegistrationForm
from .models import CustomUser, DutySchedule, ChangeLog, RegistrationRequest, ChatMessage, Notification
from django.contrib import messages
from django.contrib.auth.views import LogoutView
from django.http import HttpResponse, JsonResponse
import csv
from openpyxl import Workbook
from io import BytesIO
from django.contrib.auth.forms import UserCreationForm, UserChangeForm
import json
from django.views.decorators.csrf import csrf_exempt
from django.http import JsonResponse
from django.contrib.auth.decorators import user_passes_test
from django.shortcuts import render, get_object_or_404
import logging
from django.utils.timezone import now, timedelta
from django.db import models  # Добавлено для Q-объектов
import calendar
from django.shortcuts import get_object_or_404
from openpyxl.styles import Font, Alignment
from .models import Task
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.cell.cell import MergedCell
import openpyxl
from .models import CustomUser, DutySchedule, ChangeLog, RegistrationRequest, ChatMessage, Notification, Task, WorkAttendance
from .create_timesheet import create_timesheet
import io
import os
from django.core.paginator import Paginator, PageNotAnInteger, EmptyPage


logger = logging.getLogger(__name__)


@login_required
@user_passes_test(lambda u: not u.is_superuser, login_url='/admin-dashboard/')
def home_view(request):
    # Проверяем предстоящие дежурства
    today = now().date()
    tomorrow = today + timedelta(days=1)
    upcoming_duties = DutySchedule.objects.filter(user=request.user, date=tomorrow)

    for duty in upcoming_duties:
        message = f"Напоминание: у вас дежурство {duty.date.strftime('%d.%m.%Y')}."
        # Проверяем, не было ли уже создано уведомление для этого дежурства (проверяем по сообщению и is_read=False)
        if not Notification.objects.filter(user=request.user, message=message, is_read=False).exists():
            Notification.objects.create(user=request.user, message=message, is_read=False)

    # Загружаем только непрочитанные уведомления для пользователя
    notifications = Notification.objects.filter(user=request.user, is_read=False)

    context = {
        'notifications': notifications,  # Передаем только непрочитанные уведомления
    }
    return render(request, 'home.html', context)


def login_view(request):
    if request.method == 'POST':
        form = LoginForm(request.POST)
        username = request.POST.get('username')
        password = request.POST.get('password')
        user = authenticate(request, username=username, password=password)
        
        if user is not None:
            login(request, user)
            return redirect('home')
        else:
            form.add_error(None, "Неверное имя пользователя или пароль.")
    else:
        form = LoginForm()
    
    return render(request, 'login.html', {'form': form})


def register_view(request):
    if request.method == 'POST':
        form = RegistrationForm(request.POST)
        if form.is_valid():
            data = form.cleaned_data
            RegistrationRequest.objects.create(
                username=data['username'],
                first_name=data['first_name'],
                last_name=data['last_name'],
                email=data['email'],
                password=data['password1']
            )
            messages.success(request, "Ваш запрос отправлен на рассмотрение администратором.")
            return redirect('login')
    else:
        form = RegistrationForm()
    return render(request, 'register.html', {'form': form})


@login_required
@user_passes_test(lambda u: u.is_superuser)
def admin_dashboard(request):
    # Получаем текущие год и месяц или используем значения из GET-параметров
    current_year = int(request.GET.get('year', now().year))
    current_month = int(request.GET.get('month', now().month))

    # Словарь для перевода номеров месяцев на полные названия на русском
    month_names_ru = {
        1: "Январь", 2: "Февраль", 3: "Март", 4: "Апрель",
        5: "Май", 6: "Июнь", 7: "Июль", 8: "Август",
        9: "Сентябрь", 10: "Октябрь", 11: "Ноябрь", 12: "Декабрь"
    }

    # Генерируем дни месяца с полной датой и цифровым форматом
    days_in_month = []
    digital_dates = []  # Для хранения дат в формате "Y-m-d"
    for day in range(1, 32):
        try:
            date = datetime(current_year, current_month, day)
            formatted_date = f"{date.day:02d} {month_names_ru[date.month]} {date.year}"  # Полная дата (например, "27 Февраль 2025")
            digital_date = date.strftime("%d.%m.%Y")  # Цифровой формат (например, "2025-02-27")
            is_weekend = date.weekday() >= 5  # Проверка на выходной (5 - суббота, 6 - воскресенье)
            days_in_month.append((current_year, current_month, day, formatted_date, is_weekend))
            digital_dates.append((digital_date, is_weekend))
        except ValueError:
            break

    # Исключаем администратора из списка пользователей
    users = CustomUser.objects.exclude(is_superuser=True)

    # Форматируем имя пользователя как "Фамилия И.О."
    formatted_users = []
    for user in users:
        initials = f"{user.first_name[0]}.{user.middle_name[0]}." if user.first_name and user.middle_name else ""
        full_name = f"{user.last_name} {initials}".strip()
        formatted_users.append({
            'id': user.id,
            'full_name': full_name,
        })

    # Экспорт в XLSX
    if 'export_xlsx' in request.GET:
        return export_duty_schedule(request, formatted_users, digital_dates, current_year, current_month)

    context = {
        'users': formatted_users,
        'days_in_month': days_in_month,
        'digital_dates': digital_dates,  # Передаем цифровые даты в шаблон
        'current_year': current_year,
        'current_month': current_month,
        'current_month_name': month_names_ru[current_month],
    }
    return render(request, 'admin_dashboard.html', context)


@login_required
@user_passes_test(lambda u: u.is_superuser)
def export_duty_schedule(request, users, digital_dates, current_year, current_month):
    # Получаем актуальные данные о дежурствах через get_duties
    duties_response = get_duties(request)  # Вызываем функцию get_duties
    duties_data = duties_response.content.decode('utf-8')
    duties = json.loads(duties_data)  # Преобразуем JSON в Python-словарь

    wb = Workbook()
    ws = wb.active
    # Заголовки таблицы
    headers = ['Пользователи'] + [date for date, _ in digital_dates]  # Используем цифровые даты для заголовков
    ws.append(headers)
    # Данные пользователей
    for user in users:
        row = [user['full_name']]
        for date_str, is_weekend in digital_dates:
            key = f"{user['id']}-{date_str}"
            duty_mark = 'Д' if duties.get(key, False) else ''
            row.append(duty_mark)
        ws.append(row)
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    response = HttpResponse(buffer.read(),
                            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="duty_schedule.xlsx"'
    return response


@login_required
@user_passes_test(lambda u: u.is_superuser)
def approve_registration_requests(request):
    requests = RegistrationRequest.objects.all()
    if request.method == 'POST':
        request_id = request.POST.get('request_id')
        action = request.POST.get('action')
        if request_id and action == 'approve':
            reg_request = RegistrationRequest.objects.get(id=request_id)
            CustomUser.objects.create_user(
                username=reg_request.username,
                first_name=reg_request.first_name,
                last_name=reg_request.last_name,
                email=reg_request.email,
                password=reg_request.password
            )
            reg_request.delete()
            ChangeLog.objects.create(user=request.user, action=f"Approved registration for {reg_request.username}")
            messages.success(request, f"Регистрация пользователя {reg_request.username} успешно подтверждена.")
        elif request_id and action == 'reject':
            reg_request = RegistrationRequest.objects.get(id=request_id)
            reg_request.delete()
            ChangeLog.objects.create(user=request.user, action=f"Rejected registration for {reg_request.username}")
            messages.warning(request, f"Запрос регистрации пользователя {reg_request.username} отклонён.")
    context = {
        'requests': requests,
    }
    return render(request, 'approve_requests.html', context)


@login_required
def user_dashboard(request):
    user = request.user
    all_duties = DutySchedule.objects.filter(user=user).order_by('-date')
    
    # Пагинация дежурств, показывая 6 дежурств на странице
    paginator = Paginator(all_duties, 6)
    page_number = request.GET.get('page', 1)
    
    try:
        duties = paginator.page(page_number)
    except PageNotAnInteger:
        duties = paginator.page(1)
    except EmptyPage:
        duties = paginator.page(paginator.num_pages)

    context = {
        'duties': duties,
        'user_info': {
            'username': user.username,
            'first_name': user.first_name,
            'last_name': user.last_name,
            'email': user.email,
            'middle_name': user.middle_name,
            'birth_date': user.birth_date,
            'position': user.position,
            'rank': user.rank,
        },
    }
    return render(request, 'user_dashboard.html', context)


class CustomUserChangeForm(UserChangeForm):
    class Meta:
        model = CustomUser
        fields = ['username', 'first_name', 'last_name', 'email', 'middle_name', 'birth_date', 'position', 'rank']

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.fields['username'].disabled = True  # Поле username доступно только для чтения


@login_required
def update_profile(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            user = request.user
            # Обновляем данные пользователя
            user.first_name = data.get('first_name', user.first_name)
            user.last_name = data.get('last_name', user.last_name)
            user.email = data.get('email', user.email)
            user.middle_name = data.get('middle_name', user.middle_name)
            # Обработка даты рождения (если строка пустая, устанавливаем None)
            birth_date_str = data.get('birth_date', '')
            user.birth_date = datetime.strptime(birth_date_str, '%Y-%m-%d').date() if birth_date_str else None
            user.position = data.get('position', user.position)
            user.rank = data.get('rank', user.rank)

            # Проверяем обязательные поля
            if not user.first_name or not user.last_name or not user.email:
                return JsonResponse({'success': False, 'error': 'Заполните обязательные поля.'}, status=400)

            user.save()
            return JsonResponse({'success': True})
        except ValueError as e:
            return JsonResponse({'success': False, 'error': f'Некорректный формат даты: {e}'}, status=400)
        except Exception as e:
            logger.error(f"Error in update_profile: {e}")
            return JsonResponse({'success': False, 'error': str(e)}, status=400)
    return JsonResponse({'success': False}, status=400)


@login_required
def analytics_view(request):
    user = request.user
    current_month = now().month
    current_year = now().year

    # Количество дежурств за текущий месяц
    monthly_duties = DutySchedule.objects.filter(
        user=user,
        date__year=current_year,
        date__month=current_month
    ).count()
    monthly_hours = monthly_duties * 24  # Подсчет часов

    # Количество дежурств за текущий квартал
    quarterly_duties = DutySchedule.objects.filter(
        user=user,
        date__year=current_year,
        date__month__gte=max(1, current_month - 2),
        date__month__lte=current_month
    ).count()
    quarterly_hours = quarterly_duties * 24  # Подсчет часов

    # Количество дежурств за текущий год
    yearly_duties = DutySchedule.objects.filter(
        user=user,
        date__year=current_year
    ).count()
    yearly_hours = yearly_duties * 24  # Подсчет часов

    # Список всех дежурств пользователя (отсортирован по дате)
    all_duties = DutySchedule.objects.filter(user=user).order_by('-date')

    # Дежурства за текущий месяц
    current_month_duties = DutySchedule.objects.filter(
        user=user,
        date__year=current_year,
        date__month=current_month
    ).order_by('date')

    # Проверка дежурства на конкретную дату
    specific_date = request.GET.get('date')
    if specific_date:
        duties_on_date = DutySchedule.objects.filter(user=user, date=specific_date).exists()
    else:
        duties_on_date = None

    # Экспорт в CSV
    if 'export_csv' in request.GET:
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="duties.csv"'
        writer = csv.writer(response)
        writer.writerow(['Дата'])
        for duty in all_duties:
            writer.writerow([duty.date])
        return response

    # Экспорт в XLSX
    if 'export_xlsx' in request.GET:
        wb = Workbook()
        ws = wb.active
        ws.append(['Дата'])
        for duty in all_duties:
            ws.append([duty.date])
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        response = HttpResponse(buffer.read(),
                                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="duties.xlsx"'
        return response

    context = {
        'monthly_duties': monthly_duties,
        'quarterly_duties': quarterly_duties,
        'yearly_duties': yearly_duties,
        'monthly_hours': monthly_hours,  # Часы за месяц
        'quarterly_hours': quarterly_hours,  # Часы за квартал
        'yearly_hours': yearly_hours,  # Часы за год
        'all_duties': all_duties,  # Все дежурства пользователя
        'current_month_duties': current_month_duties,  # Дежурства за текущий месяц
        'duties_on_date': duties_on_date,
    }
    return render(request, 'analytics.html', context)


@login_required
def user_analytics_view(request):
    user = request.user
    current_month = now().month
    current_year = now().year

    # Общая статистика
    monthly_duties = DutySchedule.objects.filter(user=user, date__year=current_year, date__month=current_month).count()
    monthly_hours = monthly_duties * 24  # Оригинальные часы дежурств (по 24 часа)
    quarterly_duties = DutySchedule.objects.filter(user=user, date__year=current_year,
                                                  date__month__gte=max(1, current_month - 2),
                                                  date__month__lte=current_month).count()
    quarterly_hours = quarterly_duties * 24
    yearly_duties = DutySchedule.objects.filter(user=user, date__year=current_year).count()
    yearly_hours = yearly_duties * 24

    # Календарь дежурств за текущий месяц
    cal = calendar.monthcalendar(current_year, current_month)
    current_month_duties = DutySchedule.objects.filter(user=user, date__year=current_year, date__month=current_month)
    duty_dates = {duty.date.day: True for duty in current_month_duties}

    # Проверка дежурства на конкретную дату
    specific_date = request.GET.get('date')
    duties_on_date = DutySchedule.objects.filter(user=user, date=specific_date).exists() if specific_date else None

    # Экспорт в CSV для Google Календаря
    if 'export_csv' in request.GET:
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="duties_{current_year}_{current_month}.csv"'
        writer = csv.writer(response)
        writer.writerow(['Subject', 'Start Date', 'All Day Event'])
        for duty in current_month_duties:
            writer.writerow([f'Дежурство {user.last_name}', duty.date.strftime('%m/%d/%Y'), 'TRUE'])
        return response

    # Экспорт в XLSX
    if 'export_xlsx' in request.GET:
        wb = Workbook()
        ws = wb.active
        ws.title = f'Дежурства {current_year}-{current_month}'

        # Заголовки
        headers = ['Дата', 'Дежурство']
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')

        # Данные
        for duty in current_month_duties:
            ws.append([duty.date.strftime('%d.%m.%Y'), 'Д'])

        # Автоподбор ширины столбцов
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        response = HttpResponse(buffer.read(),
                               content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="duties_{current_year}_{current_month}.xlsx"'
        return response

    context = {
        'monthly_duties': monthly_duties,
        'quarterly_duties': quarterly_duties,
        'yearly_duties': yearly_duties,
        'monthly_hours': monthly_hours,
        'quarterly_hours': quarterly_hours,
        'yearly_hours': yearly_hours,
        'calendar': cal,
        'duty_dates': duty_dates,
        'current_month': current_month,
        'current_year': current_year,
        'duties_on_date': duties_on_date,
    }
    return render(request, 'user_analytics.html', context)


@login_required
def chat_view(request, receiver_id):
    receiver = get_object_or_404(CustomUser, id=receiver_id)

    # Ограничение: обычный пользователь может общаться только с администратором
    if not request.user.is_superuser and not receiver.is_superuser:
        messages.error(request, "Вы можете общаться только с администратором.")
        return redirect('chat_with_admin')

    messages_list = ChatMessage.objects.filter(
        models.Q(sender=request.user, receiver=receiver) |
        models.Q(sender=receiver, receiver=request.user)
    ).order_by('timestamp')

    # Отмечаем сообщения как прочитанные
    ChatMessage.objects.filter(sender=receiver, receiver=request.user, is_read=False).update(is_read=True)

    if request.method == 'POST':
        message_text = request.POST.get('message')
        if message_text:
            ChatMessage.objects.create(sender=request.user, receiver=receiver, message=message_text)
            return redirect('chat', receiver_id=receiver_id)

    context = {
        'receiver': receiver,
        'messages': messages_list,
    }
    return render(request, 'chat.html', context)


@login_required
@user_passes_test(lambda u: u.is_superuser)
def chat_list(request):
    users = CustomUser.objects.exclude(is_superuser=True)
    chat_previews = []

    for user in users:
        last_message = ChatMessage.objects.filter(
            models.Q(sender=request.user, receiver=user) |
            models.Q(sender=user, receiver=request.user)
        ).order_by('-timestamp').first()

        unread_count = ChatMessage.objects.filter(
            sender=user, receiver=request.user, is_read=False
        ).count()

        chat_previews.append({
            'user': user,
            'last_message': last_message,
            'unread_count': unread_count,
        })

    context = {
        'chat_previews': chat_previews,
    }
    return render(request, 'chat_list.html', context)


@login_required
def chat_with_admin(request):
    if request.user.is_superuser:
        return redirect('chat_list')  # Администратор перенаправляется на список чатов

    admin = CustomUser.objects.filter(is_superuser=True).first()
    if not admin:
        messages.error(request, "Администратор не найден.")
        return redirect('user_dashboard')

    return redirect('chat', receiver_id=admin.id)


class CustomLogoutView(LogoutView):
    http_method_names = ['get', 'post']

    def dispatch(self, request, *args, **kwargs):
        messages.success(request, "Вы успешно вышли из системы.")
        response = super().dispatch(request, *args, **kwargs)
        return redirect('login')  # Явно перенаправляем на страницу входа


@login_required
def update_profile(request):
    if request.method == 'POST':
        user = request.user
        data = json.loads(request.body)
        # Обновляем данные пользователя
        user.first_name = data.get('first_name', user.first_name)
        user.last_name = data.get('last_name', user.last_name)
        user.email = data.get('email', user.email)
        user.middle_name = data.get('middle_name', user.middle_name)
        user.birth_date = data.get('birth_date', user.birth_date)
        user.position = data.get('position', user.position)
        user.rank = data.get('rank', user.rank)
        user.save()
        return JsonResponse({'success': True})
    return JsonResponse({'success': False}, status=400)


@csrf_exempt
@login_required
@user_passes_test(lambda u: u.is_superuser)
def assign_duty(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            user_id = data.get('user_id')
            date = data.get('date')

            # Логирование для отладки
            print(f"Received user_id: {user_id}, date: {date}")

            # Проверяем входные данные
            if not user_id or not date:
                return JsonResponse({'success': False, 'error': 'Недостаточно данных.'}, status=400)

            # Преобразуем дату в объект datetime
            try:
                duty_date = datetime.strptime(date, "%Y-%m-%d").date()
            except ValueError as e:
                return JsonResponse({'success': False, 'error': f'Некорректный формат даты: {e}'}, status=400)

            # Проверяем, существует ли дежурство
            duty = DutySchedule.objects.filter(user_id=user_id, date=duty_date).first()
            if duty:
                duty.delete()  # Удаляем дежурство, если оно уже существует
                print(f"Deleted duty for user {user_id} on {duty_date}")  # Логирование
                return JsonResponse({'success': True, 'action': 'removed'})
            else:
                DutySchedule.objects.create(user_id=user_id, date=duty_date)  # Создаем новое дежурство
                print(f"Created duty for user {user_id} on {duty_date}")  # Логирование
                return JsonResponse({'success': True, 'action': 'assigned'})
        except Exception as e:
            print(f"Error in assign_duty: {e}")  # Логирование ошибок
            return JsonResponse({'success': False, 'error': str(e)}, status=400)
    return JsonResponse({'success': False}, status=400)


@csrf_exempt
@login_required
@user_passes_test(lambda u: u.is_superuser)
def save_duties(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            for change in data:
                user_id = change.get('user_id')
                date = change.get('date')
                action = change.get('action')  # 'add' или 'remove'

                if not user_id or not date:
                    return JsonResponse({'success': False, 'error': 'Недостаточно данных.'}, status=400)

                duty_date = datetime.strptime(date, "%Y-%m-%d").date()
                duty = DutySchedule.objects.filter(user_id=user_id, date=duty_date).first()

                if action == 'add':
                    if not duty:
                        DutySchedule.objects.create(user_id=user_id, date=duty_date)
                elif action == 'remove':
                    if duty:
                        duty.delete()

            return JsonResponse({'success': True})
        except Exception as e:
            print(f"Error in save_duties: {e}")
            return JsonResponse({'success': False, 'error': str(e)}, status=400)
    return JsonResponse({'success': False}, status=400)


@login_required
@user_passes_test(lambda u: u.is_superuser)
def admin_analytics_view(request):
    if not request.user.is_superuser:
        return render(request, '403.html', status=403)  # Или перенаправление

    users = CustomUser.objects.exclude(is_superuser=True)
    user_data = []

    current_month = now().month
    current_year = now().year

    for user in users:
        # Дежурства и часы за текущий месяц
        monthly_duties = DutySchedule.objects.filter(
            user=user,
            date__year=current_year,
            date__month=current_month
        ).count()
        monthly_hours = monthly_duties * 24  # Подсчет часов

        # Дежурства и часы за текущий квартал
        quarterly_duties = DutySchedule.objects.filter(
            user=user,
            date__year=current_year,
            date__month__gte=max(1, current_month - 2),
            date__month__lte=current_month
        ).count()
        quarterly_hours = quarterly_duties * 24  # Подсчет часов

        # Дежурства и часы за текущий год
        yearly_duties = DutySchedule.objects.filter(
            user=user,
            date__year=current_year
        ).count()
        yearly_hours = yearly_duties * 24  # Подсчет часов

        user_data.append({
            'id': user.id,
            'last_name': user.last_name,
            'first_name': user.first_name,
            'middle_name': user.middle_name,
            'monthly_duties': monthly_duties,
            'monthly_hours': monthly_hours,
            'quarterly_duties': quarterly_duties,
            'quarterly_hours': quarterly_hours,
            'yearly_duties': yearly_duties,
            'yearly_hours': yearly_hours,
        })

    # Экспорт в XLSX
    if 'export_xlsx' in request.GET:
        wb = Workbook()
        ws = wb.active
        # Заголовки таблицы
        headers = ['ФИО', 'Дежурства (месяц)', 'Часы (месяц)', 'Дежурства (квартал)', 'Часы (квартал)',
                   'Дежурства (год)', 'Часы (год)']
        ws.append(headers)
        # Данные пользователей
        for user in user_data:
            full_name = f"{user['last_name']} {user['first_name']} {user['middle_name']}".strip()
            row = [
                full_name,
                user['monthly_duties'],
                user['monthly_hours'],
                user['quarterly_duties'],
                user['quarterly_hours'],
                user['yearly_duties'],
                user['yearly_hours'],
            ]
            ws.append(row)
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        response = HttpResponse(buffer.read(),
                                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="user_analytics.xlsx"'
        return response

    context = {
        'users': user_data,
    }
    return render(request, 'admin_analytics.html', context)


@login_required
@user_passes_test(lambda u: u.is_superuser)
def get_duties(request):
    # Получаем текущие значения year и month из GET-параметров
    current_year = int(request.GET.get('year', now().year))
    current_month = int(request.GET.get('month', now().month))

    # Загружаем дежурства за указанный месяц
    duties = DutySchedule.objects.filter(date__year=current_year, date__month=current_month)

    # Преобразуем дежурства в словарь
    duties_dict = {}
    for duty in duties:
        key = f"{duty.user_id}-{duty.date.strftime('%Y-%m-%d')}"
        duties_dict[key] = True

    return JsonResponse(duties_dict)


@login_required
@user_passes_test(lambda u: u.is_superuser)
def users_list(request):
    users = CustomUser.objects.exclude(is_superuser=True).order_by('last_name')
    context = {
        'users': users,
    }
    return render(request, 'users_list.html', context)


@login_required
@user_passes_test(lambda u: u.is_superuser)
def edit_user(request, user_id):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            logger.info(f"Received data: {data}")  # Логирование входных данных
            user = CustomUser.objects.get(id=user_id)

            # Обновляем поля пользователя
            user.first_name = data.get('first_name', user.first_name)
            user.last_name = data.get('last_name', user.last_name)
            user.email = data.get('email', user.email)
            user.middle_name = data.get('middle_name', user.middle_name)
            user.birth_date = data.get('birth_date') or None
            user.position = data.get('position', user.position)
            user.rank = data.get('rank', user.rank)

            # Проверяем обязательные поля
            if not user.first_name or not user.last_name or not user.email:
                return JsonResponse({'success': False, 'error': 'Заполните обязательные поля.'}, status=400)

            user.save()
            return JsonResponse({'success': True})
        except CustomUser.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Пользователь не найден.'}, status=404)
        except Exception as e:
            logger.error(f"Error in edit_user: {e}")  # Логирование ошибок
            return JsonResponse({'success': False, 'error': str(e)}, status=400)
    return JsonResponse({'success': False, 'error': 'Неверный метод запроса.'}, status=400)


@login_required
@user_passes_test(lambda u: u.is_superuser)
def delete_user(request, user_id):
    user = get_object_or_404(CustomUser, id=user_id)
    user.delete()
    return JsonResponse({'success': True})


@login_required
@user_passes_test(lambda u: u.is_superuser)
def get_user_data(request, user_id):
    try:
        user = CustomUser.objects.get(id=user_id)
        data = {
            'id': user.id,
            'first_name': user.first_name,
            'last_name': user.last_name,
            'email': user.email,
            'middle_name': user.middle_name,
            'birth_date': user.birth_date.strftime('%Y-%m-%d') if user.birth_date else '',
            'position': user.position,
            'rank': user.rank,
        }
        return JsonResponse(data)
    except CustomUser.DoesNotExist:
        return JsonResponse({'error': 'Пользователь не найден.'}, status=404)


@login_required
def get_user_duties_by_month(request):
    user = request.user
    month = int(request.GET.get('month'))
    year = int(request.GET.get('year'))

    # Получаем дежурства за выбранный месяц
    duties = DutySchedule.objects.filter(
        user=user,
        date__year=year,
        date__month=month
    ).order_by('date')

    # Преобразуем данные в список словарей
    duties_data = [{'date': duty.date.strftime('%Y-%m-%d')} for duty in duties]

    return JsonResponse(duties_data, safe=False)


def check_upcoming_duties():
    today = now().date()
    tomorrow = today + timedelta(days=1)
    upcoming_duties = DutySchedule.objects.filter(date=tomorrow)

    for duty in upcoming_duties:
        user = duty.user

        # Проверяем, что пользователь НЕ является администратором
        if user.is_superuser:
            continue  # Пропускаем администраторов

        message = f"Напоминание: у вас дежурство {duty.date.strftime('%d.%m.%Y')}."
        if not Notification.objects.filter(user=user, message=message).exists():
            Notification.objects.create(user=user, message=message)


@login_required
def mark_notification_as_read(request, notification_id):
    if request.method == 'POST':
        notification = get_object_or_404(Notification, id=notification_id, user=request.user)
        notification.is_read = True  # Помечаем уведомление как прочитанное, вместо удаления
        notification.save()
        return JsonResponse({'success': True})
    return JsonResponse({'success': False}, status=400)

@login_required
def delete_notification(request, notification_id):
    if request.method == 'POST':
        notification = get_object_or_404(Notification, id=notification_id, user=request.user)
        notification.delete()  # Полностью удаляем уведомление из базы данных
        return JsonResponse({'success': True})
    return JsonResponse({'success': False}, status=400)

@login_required
def workday_view(request):
    if request.method == 'POST':
        if 'add_task' in request.POST:
            task_text = request.POST.get('task_text')
            if task_text:
                Task.objects.create(user=request.user, text=task_text)
        elif 'toggle_task' in request.POST:
            task_id = request.POST.get('task_id')
            task = Task.objects.get(id=task_id, user=request.user)
            task.completed = not task.completed
            task.save()
            return JsonResponse({'success': True})
        elif 'mark_attendance' in request.POST:  # Новая логика для отметки посещения
            date = now().date()
            attendance, created = WorkAttendance.objects.get_or_create(
                user=request.user,
                date=date,
                defaults={'hours_worked': 8, 'is_present': True}
            )
            if not created:
                attendance.is_present = True  # Обновляем, если запись уже существует
                attendance.save()
            messages.success(request, "Посещение отмечено.")
            return JsonResponse({'success': True})

    tasks = Task.objects.filter(user=request.user, date=now().date())

    # Экспорт в Т-13, если запрошен
    if 'export_t13' in request.GET:
        user = request.user
        current_month = now().month
        current_year = now().year

        wb = Workbook()
        ws = wb.active
        ws.title = 'Табель учета рабочего времени'

        # Настройки стилей
        bold_font = Font(bold=True)
        center_alignment = Alignment(horizontal='center', vertical='center')
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        # Заголовок формы
        ws.merge_cells('A1:F1')
        ws['A1'] = 'Табель учета рабочего времени'
        ws['A1'].font = bold_font
        ws['A1'].alignment = center_alignment

        # Информация об организации
        ws['A2'] = 'Наименование организации: УМВД России по Брянской области'
        ws['A3'] = 'Структурное подразделение: Вычислительный центр ИЦ'
        ws['A4'] = f'Период: {calendar.month_name[current_month]} {current_year}'

        # Заголовки таблицы
        headers = ['ФИО', 'Должность', 'Табельный номер', 'День', 'Отметка о работе', 'Часы']
        ws.append(headers)
        for cell in ws[6]:  # Считаем, что строка 6 — это строка с заголовками
            cell.font = bold_font
            cell.alignment = center_alignment
            cell.border = thin_border

        # Данные сотрудника
        full_name = f"{user.last_name} {user.first_name} {user.middle_name}".strip()
        position = user.position if user.position else 'Не указана'
        employee_number = user.id  # Используем ID пользователя как табельный номер (можно заменить на реальный)

        # Получаем все посещения за месяц
        attendance_days = WorkAttendance.objects.filter(
            user=user,
            date__year=current_year,
            date__month=current_month,
            is_present=True
        ).order_by('date')

        row = 7  # Начинаем с 7-й строки (после заголовков)
        ws[f'A{row}'] = full_name
        ws[f'B{row}'] = position
        ws[f'C{row}'] = employee_number

        # Заполняем таблицу днями и отметками
        for attendance in attendance_days:
            day = attendance.date.day
            ws[f'D{row}'] = f'{day:02d}.{current_month:02d}.{current_year}'
            ws[f'E{row}'] = 'Я'  # Отметка о работе (Я — явка)
            ws[f'F{row}'] = attendance.hours_worked  # Часы работы (по умолчанию 8)
            row += 1

        # Форматирование ячеек
        for r in range(7, row):  # Форматируем все строки с данными
            for col in ['A', 'B', 'C', 'D', 'E', 'F']:
                cell = ws[f'{col}{r}']
                cell.alignment = center_alignment
                cell.border = thin_border

        # Автоподбор ширины столбцов (с учетом MergedCell)
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter if not isinstance(col[0], MergedCell) else col[1].column_letter
            for cell in col:
                if not isinstance(cell, MergedCell):  # Пропускаем объединенные ячейки
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        response = HttpResponse(buffer.read(),
                               content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="tabel_{current_year}_{current_month}.xlsx"'
        return response

    context = {
        'tasks': tasks,
    }
    return render(request, 'workday.html', context)

@login_required
def toggle_task(request, task_id):
    if request.method == 'POST':
        task = get_object_or_404(Task, id=task_id, user=request.user)
        task.completed = not task.completed
        task.save()
        return JsonResponse({'success': True})
    return JsonResponse({'success': False}, status=400)

def export_attendance(request):
    if request.method == 'POST':
        start_date_str = request.POST.get('start_date')
        end_date_str = request.POST.get('end_date')
        
        try:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d')
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d')
            
            # Получаем данные о пользователе
            user_data = {
                'id': request.user.id,
                'first_name': request.user.first_name,
                'last_name': request.user.last_name,
                'middle_name': request.user.middle_name,
            }
            
            # Получаем данные о посещаемости
            attendance_data = {}
            current_date = start_date
            while current_date <= end_date:
                # Проверяем, есть ли запись о посещаемости на этот день
                attendance = WorkAttendance.objects.filter(
                    user=request.user,
                    date=current_date.date()
                ).first()
                
                attendance_data[current_date.date()] = attendance.is_present if attendance else False
                current_date += timedelta(days=1)
            
            # Создаем документ
            doc = create_timesheet(user_data, attendance_data, start_date, end_date)
            
            # Сохраняем во временный буфер
            buffer = io.BytesIO()
            doc.save(buffer)
            buffer.seek(0)
            
            # Формируем имя файла
            filename = f"Табель_{request.user.last_name}_{start_date_str}_{end_date_str}.docx"
            
            # Отправляем файл
            response = HttpResponse(
                buffer.getvalue(),
                content_type='application/vnd.openxmlformats-officedocument.wordprocessingml.document'
            )
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            return response
            
        except Exception as e:
            messages.error(request, f'Ошибка при создании табеля: {str(e)}')
            return redirect('workday')
            
    return redirect('workday')

@login_required
def add_task(request):
    if request.method == 'POST':
        description = request.POST.get('description')
        if description:
            Task.objects.create(
                user=request.user,
                description=description,
                date=now().date()
            )
            messages.success(request, 'Задача добавлена')
        else:
            messages.error(request, 'Необходимо указать описание задачи')
    return redirect('workday')

@login_required
def delete_task(request, task_id):
    if request.method == 'POST':
        task = get_object_or_404(Task, id=task_id, user=request.user)
        task.delete()
        messages.success(request, 'Задача удалена')
    return redirect('workday')

@login_required
def get_duties_json(request):
    user = request.user
    all_duties = DutySchedule.objects.filter(user=user).order_by('-date')
    
    # Пагинация дежурств, 6 дежурств на странице
    paginator = Paginator(all_duties, 6)
    page_number = request.GET.get('page', 1)
    
    try:
        duties = paginator.page(page_number)
    except PageNotAnInteger:
        duties = paginator.page(1)
    except EmptyPage:
        duties = paginator.page(paginator.num_pages)
    
    # Преобразуем данные о дежурствах в JSON
    duties_data = []
    for duty in duties:
        duties_data.append({
            'date': duty.date.strftime('%d.%m.%Y'),
            'weekday': duty.date.strftime('%A').capitalize(),
        })
    
    # Готовим данные о пагинации
    pagination_info = {
        'current_page': duties.number,
        'total_pages': duties.paginator.num_pages,
        'has_previous': duties.has_previous(),
        'has_next': duties.has_next(),
        'previous_page': duties.previous_page_number() if duties.has_previous() else None,
        'next_page': duties.next_page_number() if duties.has_next() else None,
        'page_range': list(duties.paginator.page_range),
    }
    
    return JsonResponse({
        'duties': duties_data,
        'pagination': pagination_info
    })